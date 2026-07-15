from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

EXCEL_PATH = Path(__file__).parent / "Productos.xlsx"
SHEET_NAME = "Productos"
IGV_RATE = 0.18
EXPECTED_PRODUCT_HEADERS = [
    "ID",
    "Nombre_Drive",
    "Producto_Web",
    "Color",
    "Kardex",
    "Nombre_Genesys",
    "Gama",
    "Precio",
    "stock_Ideal",
    "Precio_sinIGV",
    "Karde_EQ",
]
FALLBACK_PRODUCT_COLUMNS = ["Producto_Web", "Nombre_Drive", "Nombre_Genesys", "Producto"]


def validate_excel_file() -> None:
    if not EXCEL_PATH.exists():
        st.error(f"No se encontró el archivo {EXCEL_PATH.name} en la carpeta del proyecto.")
        st.stop()

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as exc:
        st.error(f"No se pudo abrir {EXCEL_PATH.name} como Excel: {exc}")
        st.stop()

    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        st.error(f"La hoja '{SHEET_NAME}' no existe en {EXCEL_PATH.name}.")
        st.stop()

    ws = wb[SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    wb.close()

    if not any(col in headers for col in FALLBACK_PRODUCT_COLUMNS):
        st.error(
            "El archivo Excel no tiene ninguna de las columnas de nombre de producto esperadas. "
            f"Debe incluir una de: {', '.join(FALLBACK_PRODUCT_COLUMNS)}."
        )
        st.stop()

    if "Precio" not in headers:
        st.error("El archivo Excel debe incluir la columna 'Precio' para generar facturas correctamente.")
        st.stop()


def normalize_product_name(value: str) -> str:
    return str(value or "").strip()


def get_product_name_series(df: pd.DataFrame) -> pd.Series:
    fallback_columns = [column for column in FALLBACK_PRODUCT_COLUMNS if column in df.columns]
    if not fallback_columns:
        return pd.Series([], dtype=str)

    def first_non_empty(row: pd.Series) -> str:
        for column in fallback_columns:
            value = row.get(column) if isinstance(row, pd.Series) else row[column]
            if pd.notna(value):
                candidate = str(value).strip()
                if candidate:
                    return candidate
        return ""

    return df.apply(first_non_empty, axis=1).astype(str)


@st.cache_data(show_spinner=False)
def load_products() -> pd.DataFrame:
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    except FileNotFoundError:
        raise RuntimeError(f"El archivo {EXCEL_PATH.name} no existe.")
    except ValueError as exc:
        raise RuntimeError(f"Error leyendo la hoja '{SHEET_NAME}' del Excel: {exc}")
    except Exception as exc:
        raise RuntimeError(f"No se pudo cargar el Excel: {exc}")

    for col in ["Precio", "Precio_sinIGV", "stock_Ideal"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Producto_Web" in df.columns:
        df["Producto_Web"] = df["Producto_Web"].astype(str).str.strip()

    for col in ["Nombre_Drive", "Nombre_Genesys", "Producto"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    return df


def get_headers() -> list[str]:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"La hoja '{SHEET_NAME}' no existe en {EXCEL_PATH.name}.")
    ws = wb[SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    wb.close()
    return [str(h) for h in headers if h is not None]


def get_next_id(df: pd.DataFrame) -> int:
    if "ID" not in df.columns:
        return 1
    ids = pd.to_numeric(df["ID"], errors="coerce").dropna()
    if ids.empty:
        return 1
    return int(ids.max()) + 1


def append_product_to_excel(row_map: dict[str, object]) -> None:
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME]
        headers = [cell.value for cell in ws[1]]

        new_row = [row_map.get(str(header), "") for header in headers]
        ws.append(new_row)
        wb.save(EXCEL_PATH)
    except Exception as exc:
        raise RuntimeError(f"No se pudo guardar el producto en Excel: {exc}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass



def make_invoice_dataframe(items: list[dict[str, object]]) -> pd.DataFrame:
    rows = []
    for item in items:
        cantidad = int(item["Cantidad"])
        precio = float(item["Precio Unitario"])
        subtotal = cantidad * precio
        rows.append(
            {
                "Producto": item["Producto"],
                "Cantidad": cantidad,
                "Precio Unitario": round(precio, 2),
                "Subtotal": round(subtotal, 2),
            }
        )
    return pd.DataFrame(rows)



def to_invoice_excel_bytes(invoice_df: pd.DataFrame, customer: str, invoice_number: str, issue_date: date) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        invoice_df.to_excel(writer, index=False, sheet_name="Factura")

        workbook = writer.book
        worksheet = writer.sheets["Factura"]

        money_fmt = workbook.add_format({"num_format": "S/ #,##0.00"})
        bold_fmt = workbook.add_format({"bold": True})

        last_row = len(invoice_df) + 2
        subtotal = float(invoice_df["Subtotal"].sum())
        igv = subtotal * IGV_RATE
        total = subtotal + igv

        worksheet.write("F1", "Factura Nro", bold_fmt)
        worksheet.write("G1", invoice_number)
        worksheet.write("F2", "Cliente", bold_fmt)
        worksheet.write("G2", customer)
        worksheet.write("F3", "Fecha", bold_fmt)
        worksheet.write("G3", issue_date.strftime("%d/%m/%Y"))

        worksheet.write(last_row, 2, "Subtotal", bold_fmt)
        worksheet.write_number(last_row, 3, subtotal, money_fmt)
        worksheet.write(last_row + 1, 2, "IGV (18%)", bold_fmt)
        worksheet.write_number(last_row + 1, 3, igv, money_fmt)
        worksheet.write(last_row + 2, 2, "Total", bold_fmt)
        worksheet.write_number(last_row + 2, 3, total, money_fmt)

        worksheet.set_column("A:A", 30)
        worksheet.set_column("B:D", 16)
        worksheet.set_column("F:G", 18)

    return output.getvalue()



def render_invoice_tab(products_df: pd.DataFrame) -> None:
    st.subheader("Generador de facturas")

    if products_df.empty:
        st.warning("No hay productos cargados en el Excel.")
        return

    if "invoice_items" not in st.session_state:
        st.session_state.invoice_items = []

    c1, c2, c3 = st.columns(3)
    with c1:
        customer = st.text_input("Cliente", placeholder="Nombre o razon social")
    with c2:
        invoice_number = st.text_input(
            "Numero de factura",
            value=f"F-{datetime.now().strftime('%Y%m%d-%H%M')}",
        )
    with c3:
        issue_date = st.date_input("Fecha", value=date.today())

    product_names = get_product_name_series(products_df)
    options = sorted([name for name in product_names.dropna().unique().tolist() if name])

    if not options:
        st.warning("No se encontraron nombres de producto válidos en el Excel.")
        return

    with st.form("add_item_form", clear_on_submit=True):
        f1, f2 = st.columns([3, 1])
        with f1:
            selected_product = st.selectbox("Producto", options=options)
        with f2:
            quantity = st.number_input("Cantidad", min_value=1, value=1, step=1)

        add_item = st.form_submit_button("Agregar a la factura", use_container_width=True)

    if add_item:
        normalized_name = selected_product.strip().lower()
        matched_rows = products_df[product_names.str.lower() == normalized_name]
        if matched_rows.empty:
            st.error("No se encontró el producto seleccionado en los datos.")
        else:
            product_row = matched_rows.iloc[0]
            unit_price = float(product_row.get("Precio", 0) or 0)

            st.session_state.invoice_items.append(
                {
                    "Producto": selected_product,
                    "Cantidad": int(quantity),
                    "Precio Unitario": unit_price,
                }
            )
            st.success(f"Producto agregado: {selected_product}")

    if st.button("Limpiar factura", type="secondary"):
        st.session_state.invoice_items = []
        st.rerun()

    if not st.session_state.invoice_items:
        st.info("Agrega productos para construir la factura.")
        return

    with st.expander("Revisar y editar ítems de la factura", expanded=True):
        for idx, item in enumerate(st.session_state.invoice_items):
            item_cols = st.columns([4, 1])
            item_cols[0].markdown(
                f"**{idx + 1}. {item['Producto']}** — Cantidad: {item['Cantidad']} — Precio unitario: S/ {float(item['Precio Unitario']):,.2f}"
            )
            if item_cols[1].button("Eliminar", key=f"remove_{idx}"):
                removed = st.session_state.invoice_items.pop(idx)
                st.success(f"Producto eliminado: {removed['Producto']}")
                st.rerun()

        if st.session_state.invoice_items:
            with st.form("edit_invoice_item_form", clear_on_submit=False):
                edit_options = [
                    f"{i + 1}. {item['Producto']} (x{item['Cantidad']})"
                    for i, item in enumerate(st.session_state.invoice_items)
                ]
                selected_option = st.selectbox("Item a editar", options=edit_options)
                selected_index = edit_options.index(selected_option)
                selected_item = st.session_state.invoice_items[selected_index]

                new_quantity = st.number_input(
                    "Cantidad",
                    min_value=1,
                    value=int(selected_item["Cantidad"]),
                    step=1,
                    key="edit_quantity",
                )
                new_price = st.number_input(
                    "Precio unitario",
                    min_value=0.0,
                    value=float(selected_item["Precio Unitario"]),
                    step=0.1,
                    format="%.2f",
                    key="edit_price",
                )

                update_item = st.form_submit_button("Actualizar item", use_container_width=True)

                if update_item:
                    st.session_state.invoice_items[selected_index]["Cantidad"] = int(new_quantity)
                    st.session_state.invoice_items[selected_index]["Precio Unitario"] = float(new_price)
                    st.success("Item actualizado correctamente.")
                    st.rerun()

    invoice_df = make_invoice_dataframe(st.session_state.invoice_items)
    subtotal = float(invoice_df["Subtotal"].sum())
    igv = subtotal * IGV_RATE
    total = subtotal + igv

    st.dataframe(invoice_df, use_container_width=True)

    m1, m2, m3 = st.columns(3)
    m1.metric("Subtotal", f"S/ {subtotal:,.2f}")
    m2.metric("IGV (18%)", f"S/ {igv:,.2f}")
    m3.metric("Total", f"S/ {total:,.2f}")

    csv_bytes = invoice_df.to_csv(index=False).encode("utf-8")
    xlsx_bytes = to_invoice_excel_bytes(invoice_df, customer, invoice_number, issue_date)

    d1, d2 = st.columns(2)
    d1.download_button(
        "Descargar factura (CSV)",
        data=csv_bytes,
        file_name=f"factura_{invoice_number}.csv",
        mime="text/csv",
        use_container_width=True,
    )
    d2.download_button(
        "Descargar factura (Excel)",
        data=xlsx_bytes,
        file_name=f"factura_{invoice_number}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )



def render_product_tab(products_df: pd.DataFrame) -> None:
    st.subheader("Agregar productos al Excel")
    st.caption("Los productos nuevos se guardan automaticamente en Productos.xlsx")

    headers = get_headers()
    next_id = get_next_id(products_df)

    with st.form("new_product_form", clear_on_submit=True):
        c1, c2 = st.columns(2)

        with c1:
            product_name = st.text_input("Nombre del producto", placeholder="Ej: Silla ergonomica")
            color = st.text_input("Color", placeholder="Ej: Negro")
            kardex = st.text_input("Kardex", placeholder="Codigo interno")
            gama = st.text_input("Gama", placeholder="Ej: Premium")

        with c2:
            price = st.number_input("Precio (con IGV)", min_value=0.0, value=0.0, step=1.0)
            stock_ideal = st.number_input("Stock ideal", min_value=0, value=0, step=1)
            karde_eq = st.text_input("Karde_EQ", placeholder="Codigo relacionado")

        submit = st.form_submit_button("Guardar producto", use_container_width=True)

    if submit:
        name_normalized = normalize_product_name(product_name)
        if not name_normalized:
            st.error("Debes ingresar el nombre del producto.")
            return

        existing_names = get_product_name_series(products_df).str.lower()
        if name_normalized.lower() in existing_names.values:
            st.error("Ese producto ya existe en el Excel.")
            return

        row_map = {
            "ID": next_id,
            "Nombre_Drive": name_normalized,
            "Producto_Web": name_normalized,
            "Color": color.strip(),
            "Kardex": kardex.strip(),
            "Nombre_Genesys": name_normalized,
            "Gama": gama.strip(),
            "Precio": float(price),
            "stock_Ideal": int(stock_ideal),
            "Precio_sinIGV": round(float(price) / (1 + IGV_RATE), 2) if price else 0,
            "Karde_EQ": karde_eq.strip(),
        }

        safe_row_map = {h: row_map.get(h, "") for h in headers}
        append_product_to_excel(safe_row_map)
        load_products.clear()
        st.success("Producto guardado correctamente en Productos.xlsx")
        st.rerun()

    st.markdown("### Productos actuales")
    st.dataframe(products_df, use_container_width=True)



def main() -> None:
    st.set_page_config(page_title="Facturador Streamlit", page_icon="🧾", layout="wide")
    st.title("App de Facturacion")
    st.caption("Genera facturas desde tu Excel de productos y agrega nuevos productos al mismo archivo.")

    validate_excel_file()
    try:
        products_df = load_products()
    except RuntimeError as exc:
        st.error(str(exc))
        return

    tab1, tab2 = st.tabs(["Generar factura", "Gestion de productos"])

    with tab1:
        render_invoice_tab(products_df)

    with tab2:
        render_product_tab(products_df)


if __name__ == "__main__":
    main()
